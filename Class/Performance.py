#! /usr/bin/env python
# coding: utf-8

import os
import datetime
import time
import hashlib
from openpyxl import Workbook, styles
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter
from JYTools.DB import DB
from Class import conf_dir

__author__ = 'ZhouHeng'


class PerformanceManager(object):

    date_format = "%Y-%m-%d"

    def __init__(self):
        self.db = DB(conf_path=os.path.join(conf_dir, "mysql_dms.conf"))
        self.t = "performance"
        self.t_module_related = "performance_module_related"
        self.t_members = "performance_members"

    def insert_performance(self, name, detail_info, start_time, end_time, user_name):
        data = dict(name=name, detail_info=detail_info, start_time=start_time, end_time=end_time, adder=user_name)
        data["insert_time"] = int(time.time())
        m = hashlib.md5()
        m.update(detail_info)
        data["id"] = m.hexdigest()
        l = self.db.execute_insert(self.t, data, ignore=True)
        if l > 0:
            return data
        return None

    def insert_module_related(self, month, module_no, performance_id):
        data = dict(month=month, module_no=module_no, id=performance_id)
        l = self.db.execute_insert(self.t_module_related, data, ignore=True)
        if l > 0:
            return data
        return None

    def insert_members(self, performance_id, user_name, score):
        data = dict(id=performance_id, user_name=user_name, score=score)
        l = self.db.execute_insert(self.t_members, data, ignore=True)
        if l > 0:
            return data
        return None

    @staticmethod
    def get_months(only_last=False):
        now_time = datetime.datetime.now()
        now_month = "%s%s" % (now_time.year, str(now_time.month).zfill(2))
        if now_time.month == 1:
            last_month = "%s%s" % (now_time.year - 1, 12)
        else:
            last_month = "%s%s" % (now_time.year, str(now_time.month - 1).zfill(2))
        if only_last is True:
            return [last_month]
        return [now_month, last_month]

    def get_performance(self, months=None):
        if isinstance(months, unicode) is True:
            if len(months) != 6:
                months = None
            else:
                months = [months]
        if months is None or len(months) == 0:
            months = self.get_months()
        r_cols = ["month", "module_no", "id"]
        r_items = self.db.execute_multi_select(self.t_module_related, where_value=dict(month=months), cols=r_cols,
                                               order_by=["id"])
        ids = map(lambda x: x["id"], r_items)
        if len(ids) <= 0:
            return []
        p_cols = ["id", "name", "detail_info", "start_time", "end_time"]
        p_items = self.db.execute_multi_select(self.t, where_value=dict(id=ids), cols=p_cols)
        pr_items = []
        pi = 0
        ri = 0
        while ri < len(r_items) and pi < len(p_items):
            if r_items[ri]["id"] == p_items[pi]["id"]:
                r_items[ri].update(p_items[pi])
                r_items[ri]["members"] = []
                pr_items.append(r_items[ri])
                pi += 1
                ri += 1
            elif r_items[ri]["id"] > p_items[pi]["id"]:
                pi += 1
            else:
                ri += 1
        m_cols = ["id", "user_name", "score"]
        m_items = self.db.execute_multi_select(self.t_members, where_value=dict(id=ids), cols=m_cols)
        pri = 0
        mi = 0
        while pri < len(pr_items) and mi < len(m_items):
            if pr_items[pri]["id"] == m_items[mi]["id"]:
                pr_items[pri]["members"].append(m_items[mi])
                mi += 1
            elif pr_items[pri]["id"] > m_items[mi]["id"]:
                mi += 1
            else:
                pri += 1
        return pr_items

    def export_performance(self, months, modules, user_items, save_path, user_name=None):
        row_height = 21
        data = self.get_performance(months)

        xls_data = dict()
        titles = ["月份", "开始日期", "完成日期", "任务名称"]
        current_user_index = None
        for i in range(len(user_items)):
            user_item = user_items[i]
            if user_item["user_name"] == user_name:
                current_user_index = i
            titles.append(user_item["nick_name"])
        titles.append("tower地址")
        for p_item in data:
            pl_item = []
            for key in ["month", "start_time", "end_time", "name"]:
                pl_item.append(p_item[key])
            pl_item[1] = time.strftime(self.date_format, time.localtime(pl_item[1]))
            pl_item[2] = time.strftime(self.date_format, time.localtime(pl_item[2]))
            for user_item in user_items:
                score = None
                for mem_item in p_item["members"]:
                    if mem_item["user_name"] == user_item["user_name"]:
                        score = float(mem_item["score"]) / 1000
                        break
                pl_item.append(score)
            pl_item.append(p_item["detail_info"])
            if p_item["module_no"] in xls_data:
                xls_data[p_item["module_no"]].append(pl_item)
            else:
                xls_data[p_item["module_no"]] = [titles, pl_item]
        wb = Workbook()
        col_len = len(user_items) + 5
        module_sheets = []
        for item in modules:
            if item["module_no"] not in xls_data:
                v = [titles]
            else:
                v = xls_data[item["module_no"]]
            ws = wb.create_sheet(item["module_name"].decode("utf-8"))

            v.append([])
            sum_row = [None, None, None, u"%s总计" % ws.title]
            for k in range(len(user_items)):
                c_location = dict(col_letter=get_column_letter(k + 5), row_start=2, row_end=len(v))
                c_v = "=SUM({col_letter}{row_start}:{col_letter}{row_end})".format(**c_location)
                sum_row.append(c_v)
            v.append(sum_row)
            for i in range(len(v)):
                ws.append(v[i])
                row_index = i + 1
                ws.row_dimensions[row_index].height = row_height
                for j in range(col_len):
                    cell_s = "%s%s" % (get_column_letter(j + 1), row_index)
                    ws[cell_s].alignment = styles.Alignment(horizontal='center', vertical='center')  # , wrapText=True
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 40
            ws.column_dimensions[get_column_letter(col_len)].width = 100
            # 对总计背景色标记为黄色
            for k in range(len(user_items) + 1):
                c_location = dict(col_letter=get_column_letter(k + 4), row_index=ws.max_row)
                cell_s = "{col_letter}{row_index}".format(**c_location)
                ws[cell_s].fill = styles.PatternFill(fill_type="solid", fgColor=colors.YELLOW)
            # 对当前用户得分加上红色
            if current_user_index is not None:
                for i in range(2, ws.max_row + 1):
                    c_location = dict(col_letter=get_column_letter(current_user_index + 5), row_index=i)
                    cell_s = "{col_letter}{row_index}".format(**c_location)
                    ws[cell_s].font = styles.Font(bold=True, color=colors.RED)
            module_sheets.append(ws)

        # 总计 sheet
        total_sheet = wb.active
        total_sheet.title = u"总计"
        total_sheet.append([None] + titles[4:-1])
        # 将每个模块sheet的总计引用到总计sheet
        for s_item in module_sheets:
            total_row = [s_item.title]
            for k in range(len(user_items)):
                c_location = dict(sheet=s_item.title, col_letter=get_column_letter(k + 5), row_index=s_item.max_row)
                c_v = u"={sheet}!{col_letter}{row_index}".format(**c_location)
                total_row.append(c_v)
            total_sheet.append(total_row)
        # 所有模块每个人的总计
        total_sum_row = [total_sheet.title]
        for k in range(len(user_items)):
            c_location = dict(col_letter=get_column_letter(k + 2), row_start=2, row_end=total_sheet.max_row)
            c_v = "=SUM({col_letter}{row_start}:{col_letter}{row_end})".format(**c_location)
            total_sum_row.append(c_v)
        total_sheet.append(total_sum_row)
        # 修改总计Sheet样式
        for j in range(1, len(modules) + 3):
            for i in range(1, len(user_items) + 2):
                cell_s = "%s%s" % (get_column_letter(i), j)
                total_sheet[cell_s].alignment = styles.Alignment(horizontal='center', vertical='center')
            total_sheet.row_dimensions[j].height = row_height
        wb.save(save_path)
        return save_path
